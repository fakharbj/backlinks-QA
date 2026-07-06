"""CSV/XLSX/paste parsing + column auto-mapping (PRD §8.3).

Accepts arbitrary headers and maps them to the canonical backlink fields. The
mapping is a best-effort guess the UI can override before processing.
"""

from __future__ import annotations

import csv
import io
from typing import Iterable

from charset_normalizer import from_bytes
from openpyxl import load_workbook

CANONICAL_FIELDS = [
    "source_page_url", "target_url", "expected_target_url", "expected_anchor_text",
    "expected_rel", "campaign", "vendor", "client_name", "cost", "placement_date",
    "expected_status", "notes", "tags",
    # Sheet-sourced fields (Phase 2)
    "assigned_user_label", "employee_code", "link_type", "sheet_created_date",
]

# Structured metadata for every canonical field, in CANONICAL_FIELDS order. Powers
# the mapping UI (labels, grouping, required markers, help text). ``required`` is
# True for ``source_page_url`` only — everything else is optional.
CANONICAL_FIELD_META: list[dict] = [
    {"key": "source_page_url", "label": "Source page URL", "required": True,
     "group": "core", "help": "The live page where the backlink appears (mandatory)."},
    {"key": "target_url", "label": "Target URL", "required": False,
     "group": "core", "help": "The linked-to URL; can default to the project target if omitted."},
    {"key": "expected_target_url", "label": "Expected target URL", "required": False,
     "group": "core", "help": "The exact target the link should point to for QA matching."},
    {"key": "expected_anchor_text", "label": "Expected anchor text", "required": False,
     "group": "core", "help": "The anchor text the link is expected to use."},
    {"key": "expected_rel", "label": "Expected rel", "required": False,
     "group": "core", "help": "Expected follow type (dofollow / nofollow / sponsored / ugc)."},
    {"key": "campaign", "label": "Campaign", "required": False,
     "group": "attribution", "help": "Campaign this placement belongs to."},
    {"key": "vendor", "label": "Vendor", "required": False,
     "group": "attribution", "help": "Supplier / provider that built the link."},
    {"key": "client_name", "label": "Client name", "required": False,
     "group": "attribution", "help": "Client the placement was done for."},
    {"key": "cost", "label": "Cost", "required": False,
     "group": "attribution", "help": "Amount paid for the placement."},
    {"key": "placement_date", "label": "Placement date", "required": False,
     "group": "meta", "help": "Date the link went live."},
    {"key": "expected_status", "label": "Expected status", "required": False,
     "group": "meta", "help": "Expected HTTP status of the source page."},
    {"key": "notes", "label": "Notes", "required": False,
     "group": "meta", "help": "Free-text notes / comments."},
    {"key": "tags", "label": "Tags", "required": False,
     "group": "meta", "help": "Free-text labels for filtering."},
    {"key": "assigned_user_label", "label": "Assigned user", "required": False,
     "group": "attribution", "help": "Team member the link is attributed to."},
    {"key": "employee_code", "label": "Employee code", "required": False,
     "group": "attribution", "help": "Staff/employee identifier for attribution."},
    {"key": "link_type", "label": "Link type", "required": False,
     "group": "meta", "help": "Free-text category (guest post, directory, profile, …)."},
    {"key": "sheet_created_date", "label": "Sheet date", "required": False,
     "group": "meta", "help": "Date the row was added in the source sheet."},
]

# Header synonyms → canonical field. Compared after lower/strip/space-collapse.
_SYNONYMS: dict[str, str] = {
    "source url": "source_page_url", "source": "source_page_url", "source page": "source_page_url",
    "source page url": "source_page_url", "page url": "source_page_url", "url": "source_page_url",
    "placement url": "source_page_url", "live url": "source_page_url",
    "target url": "target_url", "target": "target_url", "destination": "target_url",
    "destination url": "target_url", "link target": "target_url", "linked url": "target_url",
    "expected target": "expected_target_url", "expected target url": "expected_target_url",
    "anchor": "expected_anchor_text", "anchor text": "expected_anchor_text",
    "expected anchor": "expected_anchor_text", "anchor_text": "expected_anchor_text",
    "rel": "expected_rel", "follow type": "expected_rel", "follow": "expected_rel",
    "dofollow": "expected_rel", "rel attribute": "expected_rel", "follow status": "expected_rel",
    # "Link type" is a FREE-TEXT category in the sheets (guest post, directory, …),
    # NOT the rel attribute — map it to its own field.
    "link type": "link_type", "linktype": "link_type", "type": "link_type",
    "placement type": "link_type",
    "user": "assigned_user_label", "assigned": "assigned_user_label",
    "assigned user": "assigned_user_label", "assigned to": "assigned_user_label",
    "employee": "assigned_user_label", "team member": "assigned_user_label",
    "employee code": "employee_code", "emp code": "employee_code",
    "employee id": "employee_code", "emp id": "employee_code", "staff code": "employee_code",
    "sheet date": "sheet_created_date", "created date": "sheet_created_date",
    "entry date": "sheet_created_date", "added date": "sheet_created_date",
    "campaign": "campaign", "campaign name": "campaign",
    "vendor": "vendor", "supplier": "vendor", "provider": "vendor", "seller": "vendor",
    "client": "client_name", "client name": "client_name", "customer": "client_name",
    "cost": "cost", "price": "cost", "amount": "cost", "paid": "cost",
    "placement date": "placement_date", "date": "placement_date", "placed": "placement_date",
    "live date": "placement_date", "published": "placement_date",
    "expected status": "expected_status", "status": "expected_status",
    "notes": "notes", "note": "notes", "comment": "notes", "comments": "notes",
    "tags": "tags", "labels": "tags", "tag": "tags",
}


def _norm_header(h: str) -> str:
    return " ".join((h or "").strip().lower().replace("_", " ").split())


def auto_map(headers: Iterable[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for header in headers:
        canonical = _SYNONYMS.get(_norm_header(header))
        if canonical:
            mapping[header] = canonical
    return mapping


def auto_map_report(headers: list[str]) -> dict:
    """Auto-map ``headers`` and report which ones matched. Pure; drives the
    mapping UI's "these columns were recognised / these were not" hints."""
    mapping = auto_map(headers)
    return {
        "mapping": mapping,
        "matched": [h for h in headers if h in mapping],
        "unmatched": [h for h in headers if h not in mapping],
    }


def _decode(data: bytes) -> str:
    best = from_bytes(data).best()
    return str(best) if best is not None else data.decode("utf-8", errors="replace")


def parse_csv(data: bytes) -> tuple[list[str], list[dict[str, str]]]:
    text = _decode(data)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    headers = list(reader.fieldnames or [])
    rows = [{(k or ""): (v or "") for k, v in row.items()} for row in reader]
    return headers, rows


def parse_xlsx(data: bytes) -> tuple[list[str], list[dict[str, str]]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(header_row)]
    rows: list[dict[str, str]] = []
    for raw in rows_iter:
        if raw is None or all(c is None for c in raw):
            continue
        rows.append({headers[i]: ("" if c is None else str(c)) for i, c in enumerate(raw) if i < len(headers)})  # noqa: E501
    wb.close()
    return headers, rows


def parse_paste(text: str) -> tuple[list[str], list[dict[str, str]]]:
    """Bulk paste: one URL/line, ``source,target`` (or tab/semicolon) pairs, or
    a full CSV block. A first line made of recognized column names (and no
    URLs) is treated as a header row — every column is kept so the mapping
    sees anchor/rel/campaign/etc. instead of silently dropping them."""
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if not lines:
        return ["source_page_url", "target_url"], []
    first = [p.strip() for p in _split_pair(lines[0])]
    looks_like_header = (
        len(first) >= 2
        and bool(auto_map(first))
        and not any("://" in p or p.lower().startswith("www.") for p in first)
    )
    if looks_like_header:
        reader = csv.reader(io.StringIO("\n".join(lines)))
        parsed = [row for row in reader if any((c or "").strip() for c in row)]
        headers = [h.strip() for h in parsed[0]]
        rows = [
            {headers[i]: (c or "").strip() for i, c in enumerate(r) if i < len(headers)}
            for r in parsed[1:]
        ]
        return headers, rows
    headers = ["source_page_url", "target_url"]
    rows = []
    for line in lines:
        parts = [p.strip() for p in _split_pair(line)]
        if len(parts) >= 2:
            rows.append({"source_page_url": parts[0], "target_url": parts[1]})
        else:
            rows.append({"source_page_url": parts[0], "target_url": ""})
    return headers, rows


def _split_pair(line: str) -> list[str]:
    for sep in ("\t", ";", ",", " | ", "|"):
        if sep in line:
            return line.split(sep)
    return [line]


def apply_mapping(raw: dict[str, str], mapping: dict[str, str]) -> dict[str, str]:
    """Project a raw row onto canonical fields using ``mapping`` (header→canonical)."""
    mapped: dict[str, str] = {}
    for header, value in raw.items():
        canonical = mapping.get(header)
        if canonical:
            mapped[canonical] = (value or "").strip()
    return mapped
