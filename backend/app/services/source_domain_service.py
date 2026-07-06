"""Source main-domain aggregation + analytics (Phase 8, features 11/12/14).

``recompute`` refreshes the stored per-domain counters with a few set-based SQL
statements (scales to millions of rows in one pass), links each backlink to its
``source_domain``, and drops orphans. ``list_domains`` / ``detail`` serve the
dashboard from the stored aggregates so ratios never scan the backlink table.
"""

from __future__ import annotations

import uuid

from sqlalchemy import ColumnElement, and_, cast, func, or_, select, text
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.types import Float

from app.core.config import settings
from app.core.deps import AuthContext
from app.core.errors import NotFoundError, ValidationAppError
from app.models.backlink import BacklinkRecord
from app.models.project import Project
from app.models.source_domain import SourceDomain

# ── Set-based refresh SQL (:ws is a uuid bind param) ──────────────────────────
_UPSERT = text(
    """
    INSERT INTO source_domains (
        id, workspace_id, domain_key, grouping, backlink_count, indexed_count,
        not_indexed_count, uncertain_count, unchecked_count, dofollow_count,
        nofollow_count, duplicate_count, link_type_distribution, avg_score,
        project_count, user_count,
        qualified_count, not_qualified_count, warning_count, fail_count,
        pending_count, referring_domains_count,
        last_recomputed_at, created_at, updated_at)
    SELECT gen_random_uuid(), workspace_id, source_domain, 'registrable',
        count(*),
        count(*) FILTER (WHERE index_status = 'indexed'),
        count(*) FILTER (WHERE index_status = 'not_indexed'),
        count(*) FILTER (WHERE index_status = 'uncertain'),
        count(*) FILTER (WHERE index_status IS NULL),
        count(*) FILTER (WHERE current_rel::text = 'dofollow'),
        count(*) FILTER (WHERE current_rel::text = 'nofollow'),
        count(*) FILTER (WHERE is_duplicate),
        '{}'::jsonb,
        round(avg(score)::numeric, 1),
        count(DISTINCT project_id),
        count(DISTINCT nullif(btrim(assigned_user_label), '')),
        -- QA-outcome buckets on the EFFECTIVE status (override wins over status).
        count(*) FILTER (WHERE coalesce(override_status, status)::text = 'PASS'),
        count(*) FILTER (WHERE coalesce(override_status, status)::text <> 'PASS'),
        count(*) FILTER (WHERE coalesce(override_status, status)::text = 'WARNING'),
        count(*) FILTER (WHERE coalesce(override_status, status)::text = 'FAIL'),
        count(*) FILTER (WHERE coalesce(override_status, status)::text = 'PENDING'),
        count(DISTINCT target_domain),
        now(), now(), now()
    FROM backlink_records
    WHERE workspace_id = :ws AND source_domain IS NOT NULL AND source_domain <> ''
    GROUP BY workspace_id, source_domain
    ON CONFLICT (workspace_id, domain_key) DO UPDATE SET
        backlink_count = EXCLUDED.backlink_count,
        indexed_count = EXCLUDED.indexed_count,
        not_indexed_count = EXCLUDED.not_indexed_count,
        uncertain_count = EXCLUDED.uncertain_count,
        unchecked_count = EXCLUDED.unchecked_count,
        dofollow_count = EXCLUDED.dofollow_count,
        nofollow_count = EXCLUDED.nofollow_count,
        duplicate_count = EXCLUDED.duplicate_count,
        avg_score = EXCLUDED.avg_score,
        project_count = EXCLUDED.project_count,
        user_count = EXCLUDED.user_count,
        qualified_count = EXCLUDED.qualified_count,
        not_qualified_count = EXCLUDED.not_qualified_count,
        warning_count = EXCLUDED.warning_count,
        fail_count = EXCLUDED.fail_count,
        pending_count = EXCLUDED.pending_count,
        referring_domains_count = EXCLUDED.referring_domains_count,
        last_recomputed_at = now(),
        updated_at = now();
    """
)

_DIST = text(
    """
    UPDATE source_domains sd
    SET link_type_distribution = COALESCE(d.dist, '{}'::jsonb), updated_at = now()
    FROM (
        SELECT source_domain, jsonb_object_agg(lt, cnt) AS dist FROM (
            SELECT source_domain, COALESCE(NULLIF(btrim(link_type), ''), '(unset)') AS lt,
                   count(*) AS cnt
            FROM backlink_records
            WHERE workspace_id = :ws AND source_domain IS NOT NULL AND source_domain <> ''
            GROUP BY source_domain, lt
        ) x GROUP BY source_domain
    ) d
    WHERE sd.workspace_id = :ws AND sd.domain_key = d.source_domain;
    """
)

_LINK = text(
    """
    UPDATE backlink_records b SET source_domain_id = sd.id
    FROM source_domains sd
    WHERE sd.workspace_id = b.workspace_id AND sd.domain_key = b.source_domain
      AND b.workspace_id = :ws AND b.source_domain_id IS DISTINCT FROM sd.id;
    """
)

# Only 'derived' rows are recompute-owned; 'imported' rows (approved from a
# domain-import batch, 0029) survive with zero backlinks — they are catalog
# entries a user explicitly added.
_ORPHAN = text(
    """
    DELETE FROM source_domains sd
    WHERE sd.workspace_id = :ws AND sd.origin = 'derived' AND NOT EXISTS (
        SELECT 1 FROM backlink_records b
        WHERE b.workspace_id = sd.workspace_id AND b.source_domain = sd.domain_key);
    """
)

# ── Computed SQL expressions (percentages over the stored backlink_count) ─────
# NULLIF avoids divide-by-zero → NULL for zero-backlink catalog rows, which
# coalesce()es to 0.0. These are reused by list sort, stats, filters and rules.
def _pct_expr(numerator: ColumnElement) -> ColumnElement:
    return func.coalesce(
        cast(numerator, Float) * 100.0
        / func.nullif(cast(SourceDomain.backlink_count, Float), 0.0),
        0.0,
    )


_QUALIFIED_PCT = _pct_expr(SourceDomain.qualified_count)
_NOT_QUALIFIED_PCT = _pct_expr(SourceDomain.not_qualified_count)
_INDEXED_PCT = _pct_expr(SourceDomain.indexed_count)

_SORT_COLUMNS: dict[str, ColumnElement] = {
    "domain": SourceDomain.domain_key,
    "backlinks": SourceDomain.backlink_count,
    "avg_score": SourceDomain.avg_score,
    "duplicates": SourceDomain.duplicate_count,
    "indexed": SourceDomain.indexed_count,
    "da": SourceDomain.da,
    "pa": SourceDomain.pa,
    "spam_score": SourceDomain.spam_score,
    "semrush_as": SourceDomain.semrush_as,
    "qualified_count": SourceDomain.qualified_count,
    "referring_domains_count": SourceDomain.referring_domains_count,
    "indexed_pct": _INDEXED_PCT,
    "qualified_pct": _QUALIFIED_PCT,
    "not_qualified_pct": _NOT_QUALIFIED_PCT,
}

# Nullable metric columns → NULLS LAST on sort so blanks never lead the list.
_NULLABLE_SORTS = {"avg_score", "da", "pa", "spam_score", "semrush_as"}

# ── Whitelisted numeric filter fields → (column, is_computed_pct) ─────────────
# Each key maps to a stored column or a computed pct expression. NEVER interpolate
# user input; only these columns are ever compared, always with bound params.
_NUMERIC_FILTER_COLUMNS: dict[str, ColumnElement] = {
    "da": SourceDomain.da,
    "pa": SourceDomain.pa,
    "spam_score": SourceDomain.spam_score,
    "semrush_as": SourceDomain.semrush_as,
    "backlink_count": SourceDomain.backlink_count,
    "qualified_count": SourceDomain.qualified_count,
    "referring_domains_count": SourceDomain.referring_domains_count,
    "qualified_pct": _QUALIFIED_PCT,
    "not_qualified_pct": _NOT_QUALIFIED_PCT,
    "indexed_pct": _INDEXED_PCT,
}

# min/max query-param name → (whitelisted field, ">=" | "<=")
_RANGE_PARAMS: dict[str, tuple[str, str]] = {
    "da_min": ("da", ">="), "da_max": ("da", "<="),
    "pa_min": ("pa", ">="), "pa_max": ("pa", "<="),
    "spam_min": ("spam_score", ">="), "spam_max": ("spam_score", "<="),
    "as_min": ("semrush_as", ">="), "as_max": ("semrush_as", "<="),
    "backlinks_min": ("backlink_count", ">="), "backlinks_max": ("backlink_count", "<="),
    "qualified_min": ("qualified_count", ">="), "qualified_max": ("qualified_count", "<="),
    "referring_min": ("referring_domains_count", ">="),
    "referring_max": ("referring_domains_count", "<="),
    "qualified_pct_min": ("qualified_pct", ">="),
    "qualified_pct_max": ("qualified_pct", "<="),
    "not_qualified_pct_min": ("not_qualified_pct", ">="),
    "not_qualified_pct_max": ("not_qualified_pct", "<="),
    "indexed_pct_min": ("indexed_pct", ">="),
    "indexed_pct_max": ("indexed_pct", "<="),
}

_OPS = {">=", "<=", ">", "<", "==", "between"}


def _cmp(col: ColumnElement, op: str, value, value2=None) -> ColumnElement:
    """Build a single WHERE comparison against a whitelisted column."""
    if op == ">=":
        return col >= value
    if op == "<=":
        return col <= value
    if op == ">":
        return col > value
    if op == "<":
        return col < value
    if op == "==":
        return col == value
    if op == "between":
        return and_(col >= value, col <= value2)
    raise ValidationAppError(f"Unsupported operator: {op}")


async def recompute(db: AsyncSession, workspace_id: uuid.UUID) -> int:
    """Refresh stored source-domain aggregates for a workspace. Idempotent."""
    params = {"ws": workspace_id}
    await db.execute(_UPSERT, params)
    await db.execute(_DIST, params)
    await db.execute(_LINK, params)
    await db.execute(_ORPHAN, params)
    await db.flush()
    return int(
        (
            await db.execute(
                select(func.count())
                .select_from(SourceDomain)
                .where(SourceDomain.workspace_id == workspace_id)
            )
        ).scalar_one()
    )


def _to_dict(sd: SourceDomain) -> dict:
    total = sd.backlink_count or 0

    def pct(n: int | None) -> float:
        return round((n or 0) * 100.0 / total, 1) if total else 0.0

    return {
        "id": sd.id,
        "domain_key": sd.domain_key,
        "grouping": sd.grouping,
        "backlink_count": sd.backlink_count,
        "indexed_count": sd.indexed_count,
        "not_indexed_count": sd.not_indexed_count,
        "uncertain_count": sd.uncertain_count,
        "unchecked_count": sd.unchecked_count,
        "indexed_pct": pct(sd.indexed_count),
        "not_indexed_pct": pct(sd.not_indexed_count),
        "dofollow_count": sd.dofollow_count,
        "nofollow_count": sd.nofollow_count,
        "dofollow_pct": pct(sd.dofollow_count),
        "duplicate_count": sd.duplicate_count,
        "qualified_count": sd.qualified_count,
        "not_qualified_count": sd.not_qualified_count,
        "qualified_pct": pct(sd.qualified_count),
        "not_qualified_pct": pct(sd.not_qualified_count),
        "referring_domains_count": sd.referring_domains_count,
        "avg_score": float(sd.avg_score) if sd.avg_score is not None else None,
        "project_count": sd.project_count,
        "user_count": sd.user_count,
        "link_type_distribution": sd.link_type_distribution or {},
        "last_recomputed_at": sd.last_recomputed_at,
        "origin": sd.origin,
        "da": sd.da,
        "pa": sd.pa,
        "spam_score": sd.spam_score,
        "semrush_as": sd.semrush_as,
        "semrush_traffic": sd.semrush_traffic,
        "semrush_keywords": sd.semrush_keywords,
        "domain_age_days": sd.domain_age_days,
        "metrics_updated_at": sd.metrics_updated_at,
    }


def _num(value) -> float | int | None:
    if value is None or value == "":
        return None
    try:
        f = float(value)
    except (TypeError, ValueError) as exc:
        raise ValidationAppError(f"Filter value must be numeric, got {value!r}") from exc
    return int(f) if f.is_integer() else f


def _build_filters(
    ctx: AuthContext,
    *,
    filters: dict | None = None,
    search: str | None = None,
    origin: str | None = None,
) -> list[ColumnElement]:
    """Translate the whitelisted query params into a list of WHERE clauses.

    Every clause compares a whitelisted column against a bound param — user input
    is never interpolated into SQL. ``filters`` carries the da_min/max… range
    params (keys from ``_RANGE_PARAMS``); ``search`` is an ilike on the domain;
    ``origin`` restricts to 'derived' | 'imported'.
    """
    clauses: list[ColumnElement] = [SourceDomain.workspace_id == ctx.workspace_id]
    if search and search.strip():
        clauses.append(SourceDomain.domain_key.ilike(f"%{search.strip()}%"))
    if origin in ("derived", "imported"):
        clauses.append(SourceDomain.origin == origin)

    for param, (field, op) in _RANGE_PARAMS.items():
        raw = (filters or {}).get(param)
        val = _num(raw)
        if val is None:
            continue
        clauses.append(_cmp(_NUMERIC_FILTER_COLUMNS[field], op, val))
    return clauses


async def _project_used_domain_keys(
    db: AsyncSession, ctx: AuthContext, project_id: uuid.UUID
) -> set[str]:
    """domain_keys this project already has backlinks from (the project_view 'used'
    join, reused so project scoping stays consistent)."""
    ctx.assert_project(project_id)
    rows = (
        await db.execute(
            select(BacklinkRecord.source_domain)
            .where(
                BacklinkRecord.workspace_id == ctx.workspace_id,
                BacklinkRecord.project_id == project_id,
                BacklinkRecord.source_domain.is_not(None),
            )
            .distinct()
        )
    ).scalars().all()
    return {r for r in rows if r}


async def list_domains(
    db: AsyncSession, ctx: AuthContext, *,
    sort: str = "backlinks", order: str = "desc", search: str | None = None,
    limit: int = 200, offset: int = 0, origin: str | None = None,
    project_id: uuid.UUID | None = None, filters: dict | None = None,
) -> dict:
    """Rich, whitelisted, paginated list. Returns ``{items, total}``.

    ``filters`` holds the da_min/max… range params. ``project_id`` restricts to
    domains USED by that project (via the source_domain of its backlinks), else
    workspace-wide. Sorting/filtering are strictly whitelisted.
    """
    limit = max(1, min(int(limit), 2000))
    offset = max(0, int(offset))
    clauses = _build_filters(ctx, filters=filters, search=search, origin=origin)
    if project_id is not None:
        keys = await _project_used_domain_keys(db, ctx, project_id)
        if not keys:
            return {"items": [], "total": 0}
        clauses.append(SourceDomain.domain_key.in_(keys))

    total = int(
        (
            await db.execute(
                select(func.count()).select_from(SourceDomain).where(*clauses)
            )
        ).scalar_one()
    )

    col = _SORT_COLUMNS.get(sort, SourceDomain.backlink_count)
    if order == "asc":
        ordering = col.asc()
    else:
        ordering = col.desc()
    if sort in _NULLABLE_SORTS:
        ordering = ordering.nullslast()

    stmt = (
        select(SourceDomain)
        .where(*clauses)
        # Stable tiebreak so keyset-free offset paging is deterministic.
        .order_by(ordering, SourceDomain.domain_key.asc())
        .limit(limit)
        .offset(offset)
    )
    items = [_to_dict(sd) for sd in (await db.execute(stmt)).scalars().all()]
    return {"items": items, "total": total}


async def source_domain_stats(
    db: AsyncSession, ctx: AuthContext, *,
    search: str | None = None, origin: str | None = None,
    project_id: uuid.UUID | None = None, filters: dict | None = None,
) -> dict:
    """One set-based aggregate over the SAME filtered source-domain population."""
    clauses = _build_filters(ctx, filters=filters, search=search, origin=origin)
    if project_id is not None:
        keys = await _project_used_domain_keys(db, ctx, project_id)
        if not keys:
            return {
                "total_domains": 0, "total_backlinks": 0, "total_qualified": 0,
                "overall_qualified_pct": 0.0, "overall_indexed_pct": 0.0,
                "avg_da": None, "avg_pa": None, "avg_spam": None, "avg_as": None,
                "count_da_ge_50": 0, "count_spam_le_5": 0, "count_indexed": 0,
            }
        clauses.append(SourceDomain.domain_key.in_(keys))

    total_backlinks = func.coalesce(func.sum(SourceDomain.backlink_count), 0)
    total_qualified = func.coalesce(func.sum(SourceDomain.qualified_count), 0)
    total_indexed = func.coalesce(func.sum(SourceDomain.indexed_count), 0)
    row = (
        await db.execute(
            select(
                func.count().label("total_domains"),
                total_backlinks.label("total_backlinks"),
                total_qualified.label("total_qualified"),
                total_indexed.label("total_indexed"),
                func.avg(SourceDomain.da).label("avg_da"),
                func.avg(SourceDomain.pa).label("avg_pa"),
                func.avg(SourceDomain.spam_score).label("avg_spam"),
                func.avg(SourceDomain.semrush_as).label("avg_as"),
                func.count().filter(SourceDomain.da >= 50).label("count_da_ge_50"),
                func.count().filter(SourceDomain.spam_score <= 5).label("count_spam_le_5"),
                func.count().filter(SourceDomain.indexed_count > 0).label("count_indexed"),
            ).where(*clauses)
        )
    ).one()

    tb = int(row.total_backlinks or 0)

    def _round(v) -> float | None:
        return round(float(v), 1) if v is not None else None

    return {
        "total_domains": int(row.total_domains or 0),
        "total_backlinks": tb,
        "total_qualified": int(row.total_qualified or 0),
        "overall_qualified_pct": round(int(row.total_qualified or 0) * 100.0 / tb, 1) if tb else 0.0,
        "overall_indexed_pct": round(int(row.total_indexed or 0) * 100.0 / tb, 1) if tb else 0.0,
        "avg_da": _round(row.avg_da),
        "avg_pa": _round(row.avg_pa),
        "avg_spam": _round(row.avg_spam),
        "avg_as": _round(row.avg_as),
        "count_da_ge_50": int(row.count_da_ge_50 or 0),
        "count_spam_le_5": int(row.count_spam_le_5 or 0),
        "count_indexed": int(row.count_indexed or 0),
    }


# ── Export ────────────────────────────────────────────────────────────────────
_EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("domain", "domain_key"),
    ("backlinks", "backlink_count"),
    ("referring domains", "referring_domains_count"),
    ("indexed %", "indexed_pct"),
    ("qualified", "qualified_count"),
    ("qualified %", "qualified_pct"),
    ("not-qualified %", "not_qualified_pct"),
    ("DA", "da"),
    ("PA", "pa"),
    ("Spam", "spam_score"),
    ("AS", "semrush_as"),
    ("traffic", "semrush_traffic"),
    ("age days", "domain_age_days"),
    ("projects", "project_count"),
    ("users", "user_count"),
    ("avg score", "avg_score"),
    ("last recomputed", "last_recomputed_at"),
]


async def export_rows(
    db: AsyncSession, ctx: AuthContext, *,
    search: str | None = None, origin: str | None = None,
    project_id: uuid.UUID | None = None, filters: dict | None = None,
    sort: str = "backlinks", order: str = "desc", limit: int = 2000,
) -> tuple[list[str], list[list]]:
    """Return (headers, rows) for CSV/XLSX export — respects ALL list filters."""
    result = await list_domains(
        db, ctx, sort=sort, order=order, search=search, limit=limit,
        origin=origin, project_id=project_id, filters=filters,
    )
    headers = [label for label, _ in _EXPORT_COLUMNS]
    rows: list[list] = []
    for d in result["items"]:
        rows.append([d.get(key) for _, key in _EXPORT_COLUMNS])
    return headers, rows


async def project_view(
    db: AsyncSession, ctx: AuthContext, project_id: uuid.UUID, *, limit: int = 500
) -> dict:
    """Project-wise source-domain usage (Phase 9 P1, owner rule):
    ``used`` — domains this project already has backlinks from (with per-project
    counts); ``available`` — domains known globally in the workspace but NOT yet
    used by this project ("new source domain for this project" candidates)."""
    ctx.assert_project(project_id)
    from sqlalchemy import text as _text

    used_rows = (
        await db.execute(
            _text(
                """
                SELECT b.source_domain AS domain_key,
                       count(*) AS project_links,
                       count(*) FILTER (WHERE b.index_status = 'indexed') AS indexed,
                       round(avg(b.score) FILTER (WHERE b.score IS NOT NULL), 1) AS avg_score,
                       sd.da AS da, sd.backlink_count AS global_links
                FROM backlink_records b
                LEFT JOIN source_domains sd
                  ON sd.workspace_id = b.workspace_id AND sd.domain_key = b.source_domain
                WHERE b.workspace_id = :ws AND b.project_id = :pid AND b.source_domain IS NOT NULL
                GROUP BY b.source_domain, sd.da, sd.backlink_count
                ORDER BY project_links DESC
                LIMIT :lim
                """
            ),
            {"ws": ctx.workspace_id, "pid": project_id, "lim": limit},
        )
    ).mappings().all()

    available_rows = (
        await db.execute(
            _text(
                """
                SELECT sd.domain_key, sd.backlink_count AS global_links, sd.da AS da,
                       sd.project_count, sd.avg_score
                FROM source_domains sd
                WHERE sd.workspace_id = :ws
                  AND NOT EXISTS (
                      SELECT 1 FROM backlink_records b
                      WHERE b.project_id = :pid AND b.source_domain = sd.domain_key
                  )
                ORDER BY sd.backlink_count DESC
                LIMIT :lim
                """
            ),
            {"ws": ctx.workspace_id, "pid": project_id, "lim": limit},
        )
    ).mappings().all()

    return {
        "used": [dict(r) for r in used_rows],
        "available": [dict(r) for r in available_rows],
        "used_count": len(used_rows),
        "available_count": len(available_rows),
    }


async def detail(db: AsyncSession, ctx: AuthContext, domain_id: uuid.UUID) -> dict:
    sd = await db.get(SourceDomain, domain_id)
    if sd is None or sd.workspace_id != ctx.workspace_id:
        raise NotFoundError("Source domain not found")
    rows = (
        await db.execute(
            select(
                BacklinkRecord.id, Project.name, BacklinkRecord.source_page_url,
                BacklinkRecord.target_url, BacklinkRecord.status, BacklinkRecord.score,
                BacklinkRecord.link_type, BacklinkRecord.index_status,
                BacklinkRecord.assigned_user_label,
            )
            .outerjoin(Project, Project.id == BacklinkRecord.project_id)
            .where(BacklinkRecord.source_domain_id == sd.id)
            .order_by(BacklinkRecord.score.desc().nullslast())
            .limit(200)
        )
    ).all()
    backlinks = [
        {
            "id": bid, "project_name": pname, "source_page_url": src, "target_url": tgt,
            "status": status.value if status is not None else None, "score": score,
            "link_type": link_type, "index_status": index_status,
            "assigned_user_label": label,
        }
        for bid, pname, src, tgt, status, score, link_type, index_status, label in rows
    ]
    return {**_to_dict(sd), "backlinks": backlinks}


async def fetch_metrics(
    db: AsyncSession, ctx: AuthContext, *, force: bool = False, limit: int | None = None,
    providers: set[str] | None = None,
) -> int:
    """Fetch + store third-party metrics for the workspace's source domains.

    Processes the highest-traffic stale domains first, capped at
    ``DOMAIN_METRICS_BATCH_LIMIT`` per call (one shared HTTP client). Domain age is
    free (RDAP); Moz/Semrush populate only when their RapidAPI key is configured.
    ``providers`` scopes the fetch to specific providers (``None`` = all).
    """
    from datetime import datetime, timedelta, timezone

    import httpx

    from app.integrations import domain_metrics

    cap = limit or settings.DOMAIN_METRICS_BATCH_LIMIT
    stmt = select(SourceDomain).where(SourceDomain.workspace_id == ctx.workspace_id)
    if not force:
        cutoff = datetime.now(timezone.utc) - timedelta(days=settings.DOMAIN_METRICS_REFRESH_DAYS)
        stmt = stmt.where(
            or_(SourceDomain.metrics_updated_at.is_(None), SourceDomain.metrics_updated_at < cutoff)
        )
    stmt = stmt.order_by(SourceDomain.backlink_count.desc()).limit(cap)
    domains = list((await db.execute(stmt)).scalars().all())
    if not domains:
        return 0

    timeout = httpx.Timeout(settings.DOMAIN_METRICS_TIMEOUT_SECONDS)
    async with httpx.AsyncClient(follow_redirects=True, timeout=timeout) as client:
        for sd in domains:
            metrics = await domain_metrics.fetch_all(sd.domain_key, client, providers=providers)
            for field, value in metrics.items():
                setattr(sd, field, value)
    await db.flush()
    return len(domains)


# ── Export file builders (mirror reports_service._to_csv / _to_xlsx) ───────────
def _cell(value) -> str:
    """Render a value for a spreadsheet cell (matches reports_service._display)."""
    from datetime import datetime as _dt
    from decimal import Decimal as _Dec

    if value is None:
        return ""
    if hasattr(value, "value"):  # enums
        return str(value.value)
    if isinstance(value, _dt):
        return value.isoformat()
    if isinstance(value, _Dec):
        return str(value)
    return str(value)


def build_csv(headers: list[str], rows: list[list]) -> bytes:
    """CSV with a UTF-8 BOM (Excel-friendly), same as reports_service."""
    import csv
    import io

    buf = io.StringIO(newline="")
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([_cell(v) for v in row])
    return buf.getvalue().encode("utf-8-sig")


def build_xlsx(headers: list[str], rows: list[list], title: str = "Source Domains") -> bytes:
    """XLSX via openpyxl — same pattern as reports_service._to_xlsx (bold header
    row, frozen header, auto-filter)."""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Source Domains")
    ws.title = "".join(ch for ch in title if ch not in r"[]:*?/\\").strip()[:31] or "Sheet1"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([_cell(v) for v in row])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Saved filters (per-workspace Setting under one key) ───────────────────────
SAVED_FILTERS_KEY = "source_domain_saved_filters"


async def _saved_filters_setting(db: AsyncSession, ctx: AuthContext):
    from app.models.settings import Setting

    return (
        await db.execute(
            select(Setting).where(
                Setting.workspace_id == ctx.workspace_id, Setting.key == SAVED_FILTERS_KEY
            )
        )
    ).scalar_one_or_none()


async def list_saved_filters(db: AsyncSession, ctx: AuthContext) -> list[dict]:
    setting = await _saved_filters_setting(db, ctx)
    if setting is None:
        return []
    return list((setting.value or {}).get("filters", []))


async def upsert_saved_filter(
    db: AsyncSession, ctx: AuthContext, name: str, params: dict
) -> list[dict]:
    """Insert-or-replace one named filter (read-modify-write in one txn)."""
    from app.models.settings import Setting

    name = name.strip()
    if not name:
        raise ValidationAppError("Filter name is required")
    setting = await _saved_filters_setting(db, ctx)
    if setting is None:
        setting = Setting(workspace_id=ctx.workspace_id, key=SAVED_FILTERS_KEY, value={"filters": []})
        db.add(setting)
    existing = list((setting.value or {}).get("filters", []))
    remaining = [f for f in existing if f.get("name") != name]
    remaining.append({"name": name, "params": params or {}})
    remaining.sort(key=lambda f: f.get("name", "").lower())
    # Reassign the whole JSONB value so SQLAlchemy flushes the change.
    setting.value = {"filters": remaining}
    await db.flush()
    return remaining


async def delete_saved_filter(db: AsyncSession, ctx: AuthContext, name: str) -> list[dict]:
    setting = await _saved_filters_setting(db, ctx)
    if setting is None:
        return []
    existing = list((setting.value or {}).get("filters", []))
    remaining = [f for f in existing if f.get("name") != name]
    setting.value = {"filters": remaining}
    await db.flush()
    return remaining
