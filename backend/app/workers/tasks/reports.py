"""Asynchronous report generation tasks.

Reports are intentionally generated from the denormalized backlink table plus
the current issue table. That keeps the worker fast for common client/campaign/
vendor exports while still carrying the evidence users need.
"""

from __future__ import annotations

import csv
import io
import uuid
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from typing import Any

from sqlalchemy import Select, func, or_, select
from sqlalchemy.exc import OperationalError

from app.core.config import settings
from app.core.logging import get_logger
from app.db.session import session_scope
from app.integrations import storage
from app.models.backlink import BacklinkRecord
from app.models.crawl import BacklinkHistory, BacklinkIssue
from app.models.enums import OverallStatus, ReportFormat, ReportStatus, ReportType
from app.models.project import Campaign, Project, Vendor
from app.models.report import Report
from app.models.scoring import ScoringRuleVersion

# Summary/pivot report types → analytics group-by dimension.
_PIVOT_GROUP_BY = {
    ReportType.SOURCE_DOMAIN_SUMMARY: "source_domain",
    ReportType.LINK_TYPE_SUMMARY: "link_type",
    ReportType.USER_PERFORMANCE: "user",
}

_PIVOT_HEADERS = [
    "Group", "Total", "Avg Score", "Pass", "Warning", "Fail", "Unknown", "Review",
    "Indexed", "Not Indexed", "Nofollow", "Duplicates", "Link Missing",
]
from app.workers.celery_app import celery_app
from app.workers.runtime import run_async

log = get_logger("worker.reports")

_BACKLINK_HEADERS = [
    "Project",
    "Source URL",
    "Target URL",
    "Expected Target URL",
    "Status",
    "Score",
    "Link Found",
    "HTTP Status",
    "Final URL",
    "Rel",
    "Anchor",
    "Indexability",
    "Robots",
    "Canonical",
    "Assigned User",
    "Employee Code",
    "Link Type",
    "Source Domain",
    "Scoring Version",
    "Index Status",
    "Duplicate",
    "Global Rank",
    "Monthly Visits",
    "Domain Authority",
    "Spam Score",
    "Issues",
    "Recommendations",
    "Vendor",
    "Campaign",
    "Client",
    "Cost",
    "Last Checked",
    "Next Check",
    "Tags",
]

_HISTORY_HEADERS = [
    "Created At",
    "Project",
    "Source URL",
    "Target URL",
    "Event",
    "Severity",
    "Field",
    "Old Value",
    "New Value",
    "Score Delta",
]


async def _generate_async(report_id: uuid.UUID) -> dict:
    from app.services import batch_service

    batch_id = None
    try:
        async with session_scope() as s:
            report = await s.get(Report, report_id)
            if report is None:
                return {"error": "report not found"}
            report.status = ReportStatus.PROCESSING
            report.error = None
            meta = {
                "id": report.id,
                "workspace_id": report.workspace_id,
                "project_id": report.project_id,
                "report_type": report.report_type,
                "format": report.format,
                "title": report.title,
                "filters": report.filters or {},
            }
            batch_id = await batch_service.start(
                "report", report.workspace_id, project_id=report.project_id,
                label=f"Report — {report.title}"[:300], started_by=report.created_by,
            )
            if batch_id is not None:
                report.batch_id = batch_id

        async with session_scope() as s:
            if meta["report_type"] == ReportType.CHANGE_HISTORY:
                headers, rows = _HISTORY_HEADERS, await _history_rows(s, meta)
            elif meta["report_type"] in _PIVOT_GROUP_BY:
                headers, rows = _PIVOT_HEADERS, await _pivot_rows(s, meta)
            else:
                headers, rows = _BACKLINK_HEADERS, await _backlink_rows(s, meta)

        data, content_type = _render_file(headers, rows, meta)
        key = (
            f"{meta['workspace_id']}/{meta['id']}/"
            f"{_slug(meta['title'])}.{meta['format'].value}"
        )
        await storage.put_bytes_async(settings.S3_BUCKET_REPORTS, key, data, content_type)

        async with session_scope() as s:
            report = await s.get(Report, report_id)
            if report is not None:
                report.status = ReportStatus.COMPLETED
                report.file_key = key
                report.file_size = len(data)
                report.row_count = len(rows)
                report.completed_at = datetime.now(timezone.utc)
        await batch_service.update(
            batch_id, totals={"total": len(rows), "done": len(rows), "ok": len(rows)}
        )
        await batch_service.finish(batch_id)
        return {"rows": len(rows), "file_key": key}
    except OperationalError:
        raise
    except Exception as exc:  # noqa: BLE001
        log.error("report_generation_failed", report_id=str(report_id), error=repr(exc))
        async with session_scope() as s:
            report = await s.get(Report, report_id)
            if report is not None:
                report.status = ReportStatus.FAILED
                report.error = str(exc)[:1000]
        await batch_service.add_log(batch_id, f"Report failed: {exc}", level="error")
        await batch_service.finish(batch_id, status="failed", error=str(exc)[:500])
        return {"error": str(exc)}


async def _backlink_rows(s, meta: dict[str, Any]) -> list[dict[str, Any]]:
    stmt: Select = (
        select(
            BacklinkRecord,
            Project.name.label("project_name"),
            Vendor.name.label("vendor_name"),
            Campaign.name.label("campaign_name"),
            ScoringRuleVersion.scope.label("srv_scope"),
            ScoringRuleVersion.version.label("srv_version"),
        )
        .join(Project, Project.id == BacklinkRecord.project_id)
        .outerjoin(Vendor, Vendor.id == BacklinkRecord.vendor_id)
        .outerjoin(Campaign, Campaign.id == BacklinkRecord.campaign_id)
        .outerjoin(ScoringRuleVersion, ScoringRuleVersion.id == BacklinkRecord.scoring_rule_version_id)
        .where(BacklinkRecord.workspace_id == meta["workspace_id"])
    )
    stmt = _apply_backlink_filters(stmt, meta)
    stmt = stmt.order_by(
        Project.name.asc(), BacklinkRecord.source_domain.asc(), BacklinkRecord.id.asc()
    ).limit(_limit(meta["filters"]))

    results = (await s.execute(stmt)).all()
    ids = [row[0].id for row in results]
    issues = await _issues_by_backlink(s, ids)

    rows: list[dict[str, Any]] = []
    for bl, project_name, vendor_name, campaign_name, srv_scope, srv_version in results:
        current_issues = issues.get(bl.id, [])
        metrics = (bl.extra or {}).get("metrics") or {}
        rows.append(
            {
                "Project": project_name,
                "Source URL": bl.source_page_url,
                "Target URL": bl.target_url,
                "Expected Target URL": bl.expected_target_url,
                "Status": _value(bl.effective_status),
                "Score": bl.score,
                "Link Found": bl.link_found,
                "HTTP Status": bl.http_status,
                "Final URL": bl.final_url,
                "Rel": _value(bl.current_rel),
                "Anchor": bl.current_anchor_text,
                "Indexability": _value(bl.indexability),
                "Robots": bl.robots_status,
                "Canonical": bl.canonical_status,
                "Assigned User": bl.assigned_user_label,
                "Employee Code": bl.employee_code,
                "Link Type": bl.link_type,
                "Source Domain": bl.source_domain,
                "Scoring Version": f"{srv_scope} v{srv_version}" if srv_scope else "",
                "Index Status": bl.index_status or "unchecked",
                "Duplicate": bl.duplicate_status or "unique",
                "Global Rank": metrics.get("global_rank"),
                "Monthly Visits": metrics.get("monthly_visits"),
                "Domain Authority": metrics.get("da"),
                "Spam Score": metrics.get("spam_score"),
                "Issues": "; ".join(i["issue"] for i in current_issues),
                "Recommendations": "; ".join(
                    i["recommendation"] for i in current_issues if i["recommendation"]
                ),
                "Vendor": vendor_name,
                "Campaign": campaign_name,
                "Client": bl.client_name,
                "Cost": bl.cost,
                "Last Checked": bl.last_checked_at,
                "Next Check": bl.next_check_at,
                "Tags": ", ".join(bl.tags or []),
            }
        )
    return rows


def _apply_backlink_filters(stmt: Select, meta: dict[str, Any]) -> Select:
    filters = meta["filters"]
    if meta["project_id"] is not None:
        stmt = stmt.where(BacklinkRecord.project_id == meta["project_id"])

    if meta["report_type"] == ReportType.FAILED_LINKS:
        stmt = stmt.where(
            func.coalesce(BacklinkRecord.override_status, BacklinkRecord.status)
            == OverallStatus.FAIL
        )

    if filters.get("project_id"):
        stmt = stmt.where(BacklinkRecord.project_id == uuid.UUID(str(filters["project_id"])))
    if filters.get("vendor_id"):
        stmt = stmt.where(BacklinkRecord.vendor_id == uuid.UUID(str(filters["vendor_id"])))
    if filters.get("campaign_id"):
        stmt = stmt.where(BacklinkRecord.campaign_id == uuid.UUID(str(filters["campaign_id"])))
    if filters.get("source_domain"):
        stmt = stmt.where(BacklinkRecord.source_domain == str(filters["source_domain"]).lower())
    if filters.get("tag"):
        stmt = stmt.where(BacklinkRecord.tags.any(str(filters["tag"])))
    # Dynamic filters mirroring the analytics layer (Phase 5/6).
    if filters.get("assigned_user_label"):
        stmt = stmt.where(BacklinkRecord.assigned_user_label == str(filters["assigned_user_label"]))
    if filters.get("link_type"):
        stmt = stmt.where(BacklinkRecord.link_type == str(filters["link_type"]))
    if filters.get("link_type_id"):
        stmt = stmt.where(BacklinkRecord.link_type_id == uuid.UUID(str(filters["link_type_id"])))
    if filters.get("scoring_rule_version_id"):
        stmt = stmt.where(
            BacklinkRecord.scoring_rule_version_id == uuid.UUID(str(filters["scoring_rule_version_id"]))
        )
    if filters.get("index_status"):
        if filters["index_status"] == "unchecked":
            stmt = stmt.where(BacklinkRecord.index_status.is_(None))
        else:
            stmt = stmt.where(BacklinkRecord.index_status == str(filters["index_status"]))
    if filters.get("duplicate_status"):
        if filters["duplicate_status"] == "duplicate":
            stmt = stmt.where(BacklinkRecord.is_duplicate.is_(True))
        else:
            stmt = stmt.where(BacklinkRecord.duplicate_status == str(filters["duplicate_status"]))
    if filters.get("status"):
        stmt = stmt.where(
            func.coalesce(BacklinkRecord.override_status, BacklinkRecord.status)
            == OverallStatus(str(filters["status"]))
        )
    if filters.get("score_min") is not None:
        stmt = stmt.where(BacklinkRecord.score >= int(filters["score_min"]))
    if filters.get("score_max") is not None:
        stmt = stmt.where(BacklinkRecord.score <= int(filters["score_max"]))
    # Custom date ranges (ISO dates from the builder; inclusive end-of-day for "to").
    if filters.get("checked_from"):
        stmt = stmt.where(
            BacklinkRecord.last_checked_at >= datetime.fromisoformat(str(filters["checked_from"]))
        )
    if filters.get("checked_to"):
        stmt = stmt.where(
            BacklinkRecord.last_checked_at
            < datetime.fromisoformat(str(filters["checked_to"])) + timedelta(days=1)
        )
    if filters.get("created_from"):
        stmt = stmt.where(
            BacklinkRecord.created_at >= datetime.fromisoformat(str(filters["created_from"]))
        )
    if filters.get("created_to"):
        stmt = stmt.where(
            BacklinkRecord.created_at
            < datetime.fromisoformat(str(filters["created_to"])) + timedelta(days=1)
        )
    if filters.get("search"):
        like = f"%{str(filters['search']).strip()}%"
        stmt = stmt.where(
            or_(
                BacklinkRecord.source_page_url.ilike(like),
                BacklinkRecord.target_url.ilike(like),
                BacklinkRecord.current_anchor_text.ilike(like),
            )
        )
    return stmt


async def _issues_by_backlink(s, ids: list[uuid.UUID]) -> dict[uuid.UUID, list[dict[str, str]]]:
    if not ids:
        return {}
    grouped: dict[uuid.UUID, list[dict[str, str]]] = defaultdict(list)
    stmt = (
        select(BacklinkIssue)
        .where(BacklinkIssue.backlink_id.in_(ids))
        .order_by(BacklinkIssue.backlink_id.asc(), BacklinkIssue.severity.asc())
    )
    for issue in (await s.execute(stmt)).scalars().all():
        grouped[issue.backlink_id].append(
            {
                "issue": f"{issue.severity.value}:{issue.label}",
                "recommendation": issue.recommendation or "",
            }
        )
    return grouped


async def _pivot_rows(s, meta: dict[str, Any]) -> list[dict[str, Any]]:
    """Grouped summary rows (source domain / link type / user) reusing the analytics
    group-by engine, so a pivot report stays consistent with the Analytics desk."""
    from app.services import analytics_service

    group_by = _PIVOT_GROUP_BY[meta["report_type"]]
    filters = dict(meta["filters"] or {})
    if meta["project_id"] is not None:
        filters["project_id"] = str(meta["project_id"])

    class _Ctx:  # minimal AuthContext for the workspace-scoped analytics query
        workspace_id = meta["workspace_id"]
        allowed_project_ids = None

    groups = await analytics_service.groups(
        s, _Ctx(), filters, group_by, limit=_limit(meta["filters"])
    )
    return [
        {
            "Group": g.get("label") or g.get("key") or "",
            "Total": g.get("total"),
            "Avg Score": g.get("avg_score"),
            "Pass": g.get("pass"),
            "Warning": g.get("warning"),
            "Fail": g.get("fail"),
            "Unknown": g.get("unknown"),
            "Review": g.get("review"),
            "Indexed": g.get("indexed"),
            "Not Indexed": g.get("not_indexed"),
            "Nofollow": g.get("nofollow"),
            "Duplicates": g.get("duplicates"),
            "Link Missing": g.get("link_missing"),
        }
        for g in groups
    ]


async def _history_rows(s, meta: dict[str, Any]) -> list[dict[str, Any]]:
    stmt = (
        select(
            BacklinkHistory,
            BacklinkRecord.source_page_url,
            BacklinkRecord.target_url,
            Project.name.label("project_name"),
        )
        .join(BacklinkRecord, BacklinkRecord.id == BacklinkHistory.backlink_id)
        .join(Project, Project.id == BacklinkHistory.project_id)
        .where(BacklinkHistory.workspace_id == meta["workspace_id"])
    )
    if meta["project_id"] is not None:
        stmt = stmt.where(BacklinkHistory.project_id == meta["project_id"])
    if meta["filters"].get("event_type"):
        stmt = stmt.where(BacklinkHistory.event_type == str(meta["filters"]["event_type"]))
    stmt = stmt.order_by(BacklinkHistory.created_at.desc()).limit(_limit(meta["filters"]))

    rows: list[dict[str, Any]] = []
    for ev, source_page_url, target_url, project_name in (await s.execute(stmt)).all():
        rows.append(
            {
                "Created At": ev.created_at,
                "Project": project_name,
                "Source URL": source_page_url,
                "Target URL": target_url,
                "Event": ev.event_type.value,
                "Severity": ev.severity.value if ev.severity else "",
                "Field": ev.field,
                "Old Value": ev.old_value,
                "New Value": ev.new_value,
                "Score Delta": ev.score_delta,
            }
        )
    return rows


def _render_file(
    headers: list[str], rows: list[dict[str, Any]], meta: dict[str, Any]
) -> tuple[bytes, str]:
    fmt: ReportFormat = meta["format"]
    if fmt == ReportFormat.CSV:
        return _to_csv(headers, rows), "text/csv; charset=utf-8"
    if fmt == ReportFormat.XLSX:
        return _to_xlsx(headers, rows, str(meta["title"])), (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    if fmt == ReportFormat.PDF:
        return _to_pdf(headers, rows, str(meta["title"])), "application/pdf"
    raise ValueError(f"Unsupported report format: {fmt}")


def _to_csv(headers: list[str], rows: list[dict[str, Any]]) -> bytes:
    buf = io.StringIO(newline="")
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({h: _display(row.get(h)) for h in headers})
    return buf.getvalue().encode("utf-8-sig")


def _to_xlsx(headers: list[str], rows: list[dict[str, Any]], title: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Report")
    ws.title = _sheet_title(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([_display(row.get(h)) for h in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _to_pdf(headers: list[str], rows: list[dict[str, Any]], title: str) -> bytes:
    """Render a landscape PDF table with fpdf2.

    fpdf2 is pure-Python (no system libraries), so PDF export works on a bare
    server — unlike the previous HTML→PDF engine which needed Pango/Cairo. We
    print a focused, readable column set; the exhaustive detail lives in CSV/XLSX.
    """
    from fpdf import FPDF
    from fpdf.fonts import FontFace

    _PDF_PRIORITY = [
        "Project", "Source URL", "Target URL", "Status", "Score",
        "Link Found", "HTTP Status", "Rel", "Issues", "Last Checked",
    ]
    cols = [h for h in _PDF_PRIORITY if h in headers] or headers[:8]
    weights = {"Source URL": 3.0, "Target URL": 3.0, "Issues": 3.0, "Project": 1.6}
    col_widths = tuple(weights.get(c, 1.0) for c in cols)

    capped = rows[:8000]

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 15)
    pdf.cell(0, 9, _latin(title), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(110, 110, 110)
    note = f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} — {len(rows)} row(s)"
    if len(rows) > len(capped):
        note += f" (showing {len(capped)}; full data in CSV/XLSX)"
    pdf.cell(0, 5, _latin(note), new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(17, 17, 17)
    pdf.ln(2)

    pdf.set_font("Helvetica", "", 7)
    with pdf.table(
        col_widths=col_widths,
        text_align="LEFT",
        first_row_as_headings=True,
        headings_style=FontFace(emphasis="BOLD", fill_color=(238, 242, 238)),
        line_height=4.6,
        wrapmode="CHAR",  # break long URLs by character so cells never overflow
    ) as table:
        head = table.row()
        for col in cols:
            head.cell(_latin(col))
        for row in capped:
            tr = table.row()
            for col in cols:
                tr.cell(_latin(_display(row.get(col))))

    return bytes(pdf.output())


def _latin(value: str) -> str:
    """fpdf2 core fonts are Latin-1 only; replace anything outside it safely."""
    text = "" if value is None else str(value)
    return text.encode("latin-1", "replace").decode("latin-1")


def _display(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "value"):
        return str(value.value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def _value(value: Any) -> str | None:
    if value is None:
        return None
    return value.value if hasattr(value, "value") else str(value)


def _limit(filters: dict[str, Any]) -> int:
    raw = filters.get("limit", 100_000)
    try:
        return max(1, min(int(raw), 100_000))
    except (TypeError, ValueError):
        return 100_000


def _slug(title: str) -> str:
    # Replace runs of non-alphanumerics with a single dash (filesystem-safe).
    cleaned = "".join(ch.lower() if ch.isalnum() else " " for ch in title)
    safe = "-".join(cleaned.split())
    return safe[:80] or "report"


def _sheet_title(title: str) -> str:
    cleaned = "".join(ch for ch in title if ch not in r"[]:*?/\\").strip()
    return (cleaned or "Report")[:31]


@celery_app.task(
    name="tasks.reports.generate_report",
    bind=True,
    acks_late=True,
    max_retries=2,
    autoretry_for=(OperationalError,),
    retry_backoff=True,
)
def generate_report(self, report_id: str) -> dict:
    return run_async(_generate_async(uuid.UUID(report_id)))
