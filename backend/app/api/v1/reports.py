"""Report endpoints: create (async generation), list, status, download."""

from __future__ import annotations

import uuid

from fastapi import APIRouter, Depends, status
from starlette.responses import StreamingResponse

from app.core.config import settings
from app.core.deps import AuthContext, AuthCtx, DbSession, ReadSession, require
from app.core.errors import ValidationAppError
from app.core.rbac import Permission
from app.integrations.storage import get_bytes_async
from app.models.enums import AuditAction, ReportStatus
from app.schemas.report import ReportCreate, ReportOut
from app.services import audit_service, report_service

_REPORT_MEDIA_TYPES = {
    "pdf": "application/pdf",
    "csv": "text/csv",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}

router = APIRouter(prefix="/reports", tags=["reports"])


@router.post("", response_model=ReportOut, status_code=status.HTTP_202_ACCEPTED)
async def create_report(
    payload: ReportCreate, db: DbSession,
    ctx: AuthContext = Depends(require(Permission.EXPORT_REPORTS)),
) -> ReportOut:
    if payload.project_id is not None:
        ctx.assert_project(payload.project_id)
    report = await report_service.create_report(db, ctx, payload)
    await audit_service.record(
        db, action=AuditAction.EXPORT, actor_user_id=ctx.user.id, workspace_id=ctx.workspace_id,
        entity_type="report", entity_id=report.id, summary=f"Report {payload.report_type.value}",
    )
    await db.commit()

    from app.workers.dispatch import enqueue_report

    enqueue_report(report.id)
    return ReportOut.model_validate(report)


@router.get("", response_model=list[ReportOut])
async def list_reports(ctx: AuthCtx, db: ReadSession) -> list[ReportOut]:
    reports = await report_service.list_reports(db, ctx)
    # Resolve project names so the UI can label reports clearly (not by id).
    project_ids = {r.project_id for r in reports if r.project_id is not None}
    names: dict = {}
    if project_ids:
        from sqlalchemy import select as _select

        from app.models.project import Project

        rows = (
            await db.execute(_select(Project.id, Project.name).where(Project.id.in_(project_ids)))
        ).all()
        names = {pid: name for pid, name in rows}

    out = []
    for r in reports:
        item = ReportOut.model_validate(r)
        item.project_name = names.get(r.project_id) if r.project_id else "All projects"
        out.append(item)
    return out


@router.get("/{report_id}", response_model=ReportOut)
async def get_report(report_id: uuid.UUID, ctx: AuthCtx, db: ReadSession) -> ReportOut:
    return ReportOut.model_validate(await report_service.get_report(db, ctx, report_id))


@router.delete("/{report_id}")
async def delete_report(
    report_id: uuid.UUID, db: DbSession,
    ctx: AuthContext = Depends(require(Permission.DELETE_RECORDS)),
) -> dict:
    """Delete one report version. The stored file becomes unreferenced (no
    storage-delete API here) — harmless, and regenerating creates a fresh key."""
    report = await report_service.get_report(db, ctx, report_id)
    title = report.title
    await db.delete(report)
    await audit_service.record(
        db, action=AuditAction.DELETE, actor_user_id=ctx.user.id, workspace_id=ctx.workspace_id,
        entity_type="report", entity_id=report_id, summary=f"Deleted report {title[:120]}",
    )
    await db.commit()
    return {"message": "Report deleted"}


@router.get("/{report_id}/download")
async def download_report(report_id: uuid.UUID, ctx: AuthCtx, db: ReadSession) -> StreamingResponse:
    report = await report_service.get_report(db, ctx, report_id)
    if report.status is not ReportStatus.COMPLETED or not report.file_key:
        raise ValidationAppError("Report is not ready for download")

    try:
        data = await get_bytes_async(settings.S3_BUCKET_REPORTS, report.file_key)
    except Exception as exc:  # noqa: BLE001 — missing/unreadable object → clean 400
        raise ValidationAppError(
            "The report file could not be read — regenerate the report and try again."
        ) from exc
    fmt = report.format.value if hasattr(report.format, "value") else str(report.format)
    # HTTP headers are latin-1 only: strip non-ASCII — em dashes in report titles
    # previously crashed downloads with a 500 UnicodeEncodeError.
    safe = (report.title or "report").strip().replace("/", "-").replace(" ", "_")
    safe = safe.encode("ascii", "ignore").decode("ascii").strip("_-. ") or "report"
    return StreamingResponse(
        iter([data]),
        media_type=_REPORT_MEDIA_TYPES.get(fmt, "application/octet-stream"),
        headers={"Content-Disposition": f'attachment; filename="{safe}.{fmt}"'},
    )


@router.get("/{report_id}/rows")
async def report_rows(
    report_id: uuid.UUID, ctx: AuthCtx, db: ReadSession,
    offset: int = 0, limit: int = 100,
) -> dict:
    """View a completed report inside the app (paginated) — no download needed.
    Parses the stored CSV/XLSX; PDFs are download-only."""
    import csv as _csv
    import io as _io

    report = await report_service.get_report(db, ctx, report_id)
    if report.status is not ReportStatus.COMPLETED or not report.file_key:
        raise ValidationAppError("Report is not ready yet")
    fmt = report.format.value if hasattr(report.format, "value") else str(report.format)
    if fmt == "pdf":
        raise ValidationAppError("PDF reports can't be shown as a table — download instead.")
    try:
        data = await get_bytes_async(settings.S3_BUCKET_REPORTS, report.file_key)
    except Exception as exc:  # noqa: BLE001
        raise ValidationAppError(
            "The report file could not be read — regenerate the report and try again."
        ) from exc

    offset = max(0, offset)
    limit = max(1, min(limit, 500))
    headers: list[str] = []
    rows: list[list[str]] = []
    total = 0
    if fmt == "csv":
        reader = _csv.reader(_io.StringIO(data.decode("utf-8-sig", errors="replace")))
        for i, row in enumerate(reader):
            if i == 0:
                headers = row
                continue
            total += 1
            if offset < total <= offset + limit:
                rows.append(row)
    else:  # xlsx
        from openpyxl import load_workbook

        wb = load_workbook(_io.BytesIO(data), read_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            vals = ["" if c is None else str(c) for c in row]
            if i == 0:
                headers = vals
                continue
            total += 1
            if offset < total <= offset + limit:
                rows.append(vals)
        wb.close()
    return {"headers": headers, "rows": rows, "total": total, "offset": offset}
